"""Parses config/taxonomy_source/Dispute_codes.xlsx — the AUTHORITATIVE
source for the network-code mapping.

Three sheets are parsed:
  1. "Dispute Taxonomy"       — 36 internal types + descriptions + evidence
  2. "Evidence Weight matrix" — 10 evidence dimensions x 10 weight profiles
  3. "Amex Code mapping"      — Amex network reason codes -> internal codes

Nothing is dropped, renamed or "fixed": codes that exist only in the Excel
(e.g. BA-09, AR-04, AR-05, Process/Retrieval entries) are preserved verbatim
and surfaced by validation/validate_taxonomy.py.
"""
from __future__ import annotations

import json
import re
from functools import lru_cache

from openpyxl import load_workbook

from backend.paths import TAXONOMY_XLSX, GENERATED_CONFIG_DIR

DIMENSION_KEYS = {
    "Delivery / Shipping Proof": "delivery_proof",
    "Product Condition Evidence": "product_condition",
    "Transaction Pattern Analysis": "transaction_pattern",
    "Merchant Documentation": "merchant_documentation",
    "Cardholder Documentation": "cardholder_documentation",
    "Communication Records": "communication_records",
    "Historical Dispute Pattern": "historical_pattern",
    "Policy / ToS Compliance": "policy_compliance",
    "Digital Access / Usage Logs": "digital_access_logs",
    "Image / Visual Analysis": "image_visual_analysis",
}

PROFILE_KEYS = {
    "NR (Not Received)": "NR",
    "QD (Quality/Desc)": "QD",
    "BA-Dup (Duplicate)": "BA_DUP",
    "BA-Amt (Amount)": "BA_AMT",
    "CR-Ret (Return/Cancel)": "CR_RET",
    "CR-Sub (Subscription)": "CR_SUB",
    "AR (Authorization)": "AR",
    "SP-Rental (Vehicle/Hotel)": "SP_RENTAL",
    "SP-Travel (Airline)": "SP_TRAVEL",
    "SP-Digital (Digital Goods)": "SP_DIGITAL",
}


def _fmt_code(v) -> str:
    """4554.0 -> '4554'; keep strings as-is."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _split_amex_codes(raw: str) -> list[str]:
    """'4512, 4534' -> ['4512','4534'];  'TBD/N/A' -> []"""
    raw = _fmt_code(raw)
    if not raw or "tbd" in raw.lower() or raw.lower() in {"n/a", "na"}:
        return []
    return [c.strip() for c in re.split(r"[,/]", raw) if c.strip().isdigit()]


def _split_visa_mc(raw: str) -> tuple[list[str], list[str]]:
    """'Visa 13.1, MC 4853/4855' -> (['13.1'], ['4853','4855'])"""
    visa, mc = [], []
    if not raw:
        return visa, mc
    for part in str(raw).split(","):
        part = part.strip()
        m = re.match(r"(?i)visa\s+(.+)", part)
        if m:
            visa += [c.strip() for c in re.split(r"[/]", m.group(1)) if c.strip()]
            continue
        m = re.match(r"(?i)mc\s+(.+)", part)
        if m:
            mc += [c.strip() for c in re.split(r"[/]", m.group(1)) if c.strip()]
    return visa, mc


def parse_taxonomy_sheet(wb) -> dict:
    ws = wb["Dispute Taxonomy"]
    types, section = {}, None
    for row in ws.iter_rows(min_row=2, values_only=True):
        first, code = row[0], row[1]
        if isinstance(first, str) and code is None:
            section = first.strip()
            continue
        if not code:
            continue
        visa, mc = _split_visa_mc(row[7])
        types[str(code).strip()] = {
            "code": str(code).strip(),
            "name": (row[2] or "").strip(),
            "description": (row[3] or "").strip(),
            "key_evidence": (row[4] or "").strip(),
            "auto_fetch_sources": (row[5] or "").strip(),
            "amex_codes": _split_amex_codes(row[6]),
            "visa_codes": visa,
            "mc_codes": mc,
            "section": section,
        }
    return types


def parse_weight_matrix(wb) -> dict:
    ws = wb["Evidence Weight matrix"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    profiles = {PROFILE_KEYS.get(h, h): {} for h in header[1:] if h}
    profile_order = [PROFILE_KEYS.get(h, h) for h in header[1:] if h]
    for row in rows[1:]:
        dim = DIMENSION_KEYS.get((row[0] or "").strip())
        if not dim:
            continue
        for pname, value in zip(profile_order, row[1:]):
            profiles[pname][dim] = float(value or 0.0)
    return profiles


def parse_amex_mapping(wb) -> list[dict]:
    ws = wb["Amex Code mapping"]
    mappings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        raw_internal = (row[3] or "").strip() if row[3] else ""
        internal_codes = []
        if raw_internal and not raw_internal.lower().startswith("n/a"):
            internal_codes = [c.strip() for c in raw_internal.split("/") if c.strip()]
        mappings.append({
            "network": "amex",
            "network_code": _fmt_code(row[0]),
            "network_description": (row[1] or "").strip(),
            "excel_category": (row[2] or "").strip(),
            "internal_codes": internal_codes,
            "raw_internal_field": raw_internal,
            "resolution_approach": (row[4] or "").strip(),
        })
    return mappings


@lru_cache(maxsize=1)
def load_excel_taxonomy() -> dict:
    wb = load_workbook(TAXONOMY_XLSX, read_only=True, data_only=True)
    parsed = {
        "types": parse_taxonomy_sheet(wb),
        "weight_profiles": parse_weight_matrix(wb),
        "amex_mappings": parse_amex_mapping(wb),
        "source_file": str(TAXONOMY_XLSX),
    }
    out = GENERATED_CONFIG_DIR / "taxonomy_from_excel.json"
    out.write_text(json.dumps(parsed, indent=2))
    return parsed


def excel_internal_codes() -> set[str]:
    """Every internal code referenced anywhere in the Excel."""
    tax = load_excel_taxonomy()
    codes = set(tax["types"].keys())
    for m in tax["amex_mappings"]:
        codes.update(m["internal_codes"])
    return codes
